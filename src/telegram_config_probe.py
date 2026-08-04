"""Read-only, value-safe comparison of Telegram settings and private workbook cells."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

_TOKEN = re.compile(r"(?<![A-Za-z0-9_-])(\d{6,}:[A-Za-z0-9_-]{20,})(?![A-Za-z0-9_-])")
_ENV_KEYS = {"TELEGRAM_BOT_TOKEN", "TELEGRAM_CHAT_ID"}


def _read_telegram_env(path: Path) -> tuple[dict[str, str], list[str]]:
    if not path.is_file():
        return {}, ["env_file_missing"]
    values: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        key, separator, value = line.partition("=")
        if separator and key.strip() in _ENV_KEYS:
            values[key.strip()] = value.strip().strip('"').strip("'")
    reasons = []
    if not values.get("TELEGRAM_BOT_TOKEN"):
        reasons.append("env_token_missing")
    if not values.get("TELEGRAM_CHAT_ID"):
        reasons.append("env_chat_id_missing")
    return values, reasons


def _read_workbook_cells(path: Path) -> tuple[dict[str, str], list[str]]:
    if not path.is_file():
        return {}, ["workbook_file_missing"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Telegrams" not in workbook.sheetnames:
            return {}, ["workbook_telegrams_sheet_missing"]
        sheet = workbook["Telegrams"]
        chat_id = str(sheet["A2"].value or "").strip()
        token_text = str(sheet["B8"].value or "")
    finally:
        workbook.close()
    reasons = []
    values = {"chat_id": chat_id, "token": ""}
    token_candidates = set(_TOKEN.findall(token_text))
    if not token_text.strip():
        reasons.append("workbook_token_missing")
    elif len(token_candidates) == 1:
        values["token"] = token_candidates.pop()
    elif len(token_candidates) > 1:
        reasons.append("workbook_token_ambiguous")
    else:
        reasons.append("workbook_token_invalid")
    if not values["chat_id"]:
        reasons.append("workbook_chat_id_missing")
    return values, reasons


def probe_telegram_config(env_path: str | Path, workbook_path: str | Path) -> dict[str, object]:
    """Compare only `.env` Telegram keys with `Telegrams!A2` and `Telegrams!B8`."""
    env, reasons = _read_telegram_env(Path(env_path))
    workbook, workbook_reasons = _read_workbook_cells(Path(workbook_path))
    reasons.extend(workbook_reasons)
    env_configured = not any(reason.startswith("env_") for reason in reasons)
    workbook_configured = not any(reason.startswith("workbook_") for reason in reasons)
    token_match = env_configured and workbook_configured and env["TELEGRAM_BOT_TOKEN"] == workbook["token"]
    chat_id_match = env_configured and workbook_configured and env["TELEGRAM_CHAT_ID"] == workbook["chat_id"]
    if env_configured and workbook_configured:
        if token_match and chat_id_match:
            role = "same_credentials"
        elif not token_match:
            role = "different_bot"
            reasons.append("token_mismatch")
        else:
            role = "different_chat"
            reasons.append("chat_id_mismatch")
    else:
        role = "incomplete"
    return {
        "env_configured": env_configured,
        "workbook_configured": workbook_configured,
        "token_match": token_match,
        "chat_id_match": chat_id_match,
        "role": role,
        "reason_codes": reasons,
    }
