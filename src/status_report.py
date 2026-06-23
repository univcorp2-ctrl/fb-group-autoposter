"""Posting-status aggregation for EstateBoard broker-OK properties.

Joins the EstateBoard export (the universe of 仲介回しOK / 中間マージンの足OK
properties) with the posting record in jobs.db, so every property gets a single
at-a-glance status: 投稿済 / 失敗 / 未投稿. Pure functions here are reused by the
Excel builder, the Notion sync, and tests.
"""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from src.estateboard_adapter import (
    PROPERTY_TYPE_JA,
    _get,
    build_location,
    format_price_yen,
    format_yield,
    is_postable,
)

STATUS_POSTED = "✅ 投稿済"
STATUS_UNCERTAIN = "⚠️ 未確認（要確認）"
STATUS_FAILED = "⚠️ 失敗"
STATUS_UNPOSTED = "⬜ 未投稿"

# Sort so verified history (posted) and actionable items (未確認/失敗) float to
# the top, then the big unposted backlog newest-first. 投稿済 means VERIFIED:
# the post was actually found live on the group (permalink). 未確認 = we clicked
# post but could not confirm it went public (e.g. approval-gated group).
_STATUS_ORDER = {STATUS_POSTED: 0, STATUS_UNCERTAIN: 1, STATUS_FAILED: 2, STATUS_UNPOSTED: 3}


@dataclass(frozen=True)
class PostRecord:
    status: str  # job_targets.status of the latest relevant row
    posted_at: str | None = None
    screenshot: str | None = None
    permalink: str | None = None
    body: str | None = None


@dataclass(frozen=True)
class PropertyStatus:
    property_id: str
    status: str
    title: str
    property_type: str
    price: str
    yield_pct: str
    location: str
    posted_at_jst: str
    screenshot: str
    permalink: str
    detail_url: str
    created_at: str
    body_preview: str
    _sort_key: tuple = field(default=(), repr=False)


def load_post_records(db_path: str) -> dict[str, PostRecord]:
    """property_id -> best PostRecord (posted/uncertain win over failed/skipped)."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT j.property_id AS pid, t.status, t.posted_at, t.screenshot, t.permalink, t.body
        FROM job_targets t JOIN jobs j ON j.job_id = t.job_id
        ORDER BY COALESCE(t.posted_at, '') ASC
        """
    ).fetchall()
    conn.close()

    best: dict[str, PostRecord] = {}
    # Verified 'posted' outranks 'uncertain' so a confirmed post always wins.
    rank = {"posted": 4, "uncertain": 3, "failed": 2, "skipped": 1, "pending": 0}
    for r in rows:
        rec = PostRecord(r["status"], r["posted_at"], r["screenshot"], r["permalink"], r["body"])
        cur = best.get(r["pid"])
        if cur is None or rank.get(rec.status, 0) >= rank.get(cur.status, 0):
            best[r["pid"]] = rec
    return best


def classify(record: PostRecord | None) -> str:
    # Only a VERIFIED post (found live on the group) counts as 投稿済. 'uncertain'
    # means submitted-but-unconfirmed and must NOT masquerade as posted.
    if record and record.status == "posted":
        return STATUS_POSTED
    if record and record.status == "uncertain":
        return STATUS_UNCERTAIN
    if record and record.status == "failed":
        return STATUS_FAILED
    return STATUS_UNPOSTED


def _to_jst(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso)
        from zoneinfo import ZoneInfo

        return dt.astimezone(ZoneInfo("Asia/Tokyo")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return iso


def build_rows(items: list[dict[str, Any]], records: dict[str, PostRecord]) -> list[PropertyStatus]:
    """One PropertyStatus per broker-OK postable property, sorted for clarity."""
    out: list[PropertyStatus] = []
    for it in items:
        if not is_postable(it):
            continue
        pid = f"eb-{it.get('propertyId') or it.get('id')}"
        rec = records.get(pid)
        status = classify(rec)
        created = str(it.get("createdAt") or "")
        body = (rec.body if rec else "") or ""
        out.append(
            PropertyStatus(
                property_id=pid,
                status=status,
                title=str(_get(it, "label") or it.get("label") or ""),
                property_type=PROPERTY_TYPE_JA.get(str(_get(it, "propertyType") or ""), ""),
                price=format_price_yen(_get(it, "price")),
                yield_pct=format_yield(_get(it, "yieldAssumedFullOccupancy")),
                location=build_location(it),
                posted_at_jst=_to_jst(rec.posted_at if rec else None),
                screenshot=(rec.screenshot if rec else "") or "",
                permalink=(rec.permalink if rec else "") or "",
                detail_url=str(it.get("detail_url") or ""),
                created_at=created[:10],
                body_preview=body.replace("\n", " ")[:60],
                _sort_key=(_STATUS_ORDER.get(status, 9), _neg_str(rec.posted_at if rec else None), _neg_str(created)),
            )
        )
    out.sort(key=lambda r: r._sort_key)
    return out


def _neg_str(s: str | None) -> str:
    """Sort helper: newer ISO strings first (descending) via inverted chars."""
    if not s:
        return "￿"  # empty sorts last
    # invert each char so lexicographic ascending == chronological descending
    return "".join(chr(0x10FFFF - ord(c)) if ord(c) < 0x10FFFF else c for c in s)


def summarize(rows: list[PropertyStatus]) -> dict[str, int]:
    counts = {STATUS_POSTED: 0, STATUS_UNCERTAIN: 0, STATUS_FAILED: 0, STATUS_UNPOSTED: 0}
    for r in rows:
        counts[r.status] = counts.get(r.status, 0) + 1
    counts["total"] = len(rows)
    return counts


# --------------------------------------------------------------------------- #
# Excel / CSV output (used by scripts/build_status_db.py and the pipeline)
# --------------------------------------------------------------------------- #
# (header, attribute, column width)
COLUMNS: list[tuple[str, str, int]] = [
    ("投稿状況", "status", 12),
    ("物件ID", "property_id", 12),
    ("物件名", "title", 34),
    ("種別", "property_type", 16),
    ("価格", "price", 12),
    ("利回り", "yield_pct", 9),
    ("所在地", "location", 30),
    ("投稿日時(JST)", "posted_at_jst", 17),
    ("保存先(スクショ)", "screenshot", 46),
    ("投稿リンク", "permalink", 24),
    ("EstateBoard", "detail_url", 30),
    ("掲載日", "created_at", 11),
    ("本文プレビュー", "body_preview", 40),
]
_FILL = {STATUS_POSTED: "C6EFCE", STATUS_UNCERTAIN: "FFEB9C", STATUS_FAILED: "FFC7CE", STATUS_UNPOSTED: "F2F2F2"}
_FONT = {STATUS_POSTED: "006100", STATUS_UNCERTAIN: "9C6500", STATUS_FAILED: "9C0006", STATUS_UNPOSTED: "808080"}


def _abs_path(root: Any, path: str) -> str:
    from pathlib import Path

    if not path:
        return ""
    p = Path(path)
    return str(p if p.is_absolute() else (Path(root) / p))


def _cell(r: PropertyStatus, attr: str, root: Any) -> Any:
    val = getattr(r, attr)
    return _abs_path(root, val) if attr == "screenshot" else val


def write_excel(rows: list[PropertyStatus], out_path: Any, *, root: Any) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "投稿管理"
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append([h for h, _, _ in COLUMNS])
    for col, (_h, _a, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    for r in rows:
        ws.append([_cell(r, attr, root) for _h, attr, _w in COLUMNS])
        sc = ws.cell(row=ws.max_row, column=1)
        sc.fill = PatternFill("solid", fgColor=_FILL.get(r.status, "FFFFFF"))
        sc.font = Font(bold=True, color=_FONT.get(r.status, "000000"))
        sc.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    wb.save(out_path)


def write_csv(rows: list[PropertyStatus], out_path: Any, *, root: Any) -> None:
    import csv

    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([h for h, _, _ in COLUMNS])
        for r in rows:
            w.writerow([_cell(r, attr, root) for _h, attr, _w in COLUMNS])


def build_status_files(source_path: Any, db_path: str, out_dir: Any, *, root: Any) -> dict[str, Any]:
    """Build posting_status.xlsx + .csv from the EstateBoard export and jobs.db."""
    import json
    from pathlib import Path

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    data = json.loads(Path(source_path).read_text(encoding="utf-8"))
    items = data.get("items", data) if isinstance(data, dict) else data
    rows = build_rows(items, load_post_records(db_path))
    xlsx, csv_path = out_dir / "posting_status.xlsx", out_dir / "posting_status.csv"
    write_excel(rows, xlsx, root=root)
    write_csv(rows, csv_path, root=root)
    return {"counts": summarize(rows), "excel": str(xlsx), "csv": str(csv_path), "rows": rows}
